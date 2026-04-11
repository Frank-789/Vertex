"""
京东商品评论爬虫 - 最终修复版（已优化）
修改内容:
1. 添加商品名称提取(进入详情页后第一时间提取)
2. 修复评论区下滑逻辑(点击全部评价后直接下滑,不移动鼠标)
3. 优化点赞数提取(过滤"回复"文字)
4. 重写表格写入逻辑(参考YouTube代码,同时写入CSV和XLSX,格式化)
5. 【新增】修复弹窗下滑操作，使用更明显的滚动方式
6. 【新增】修改商品间延时为30-50秒，直接返回列表页不关闭弹窗
7. 【新增】调整字段顺序并新增"满意度"字段（超赞/一般/无随机填充）
8. 【最新修改】评论区无限下滑，连续5次无新数据才停止
9. 【最新修改】添加拟人化headers配置
"""

from DrissionPage import ChromiumPage, ChromiumOptions
from DrissionPage.errors import ElementNotFoundError, ElementLostError, NoRectError
import time
import random
import re
import csv
import logging
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('jd_scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class JDCommentScraper:
    """京东评论爬虫"""
    
    def __init__(self, keyword):
        """初始化"""
        self.keyword = keyword
        self.base_url = "https://www.jd.com/"
        
        # 持久化配置
        self.user_data_dir = os.path.join(os.getcwd(), "jd_chrome_profile")
        if not os.path.exists(self.user_data_dir):
            os.makedirs(self.user_data_dir)
            logging.info(f"✅ 创建持久化配置目录: {self.user_data_dir}")
        
        # 初始化浏览器
        self.init_browser()
        
        # CSV和Excel文件
        self.csv_file = None
        self.csv_writer = None
        self.csv_f = None
        self.excel_file = None
        self.wb = None
        self.ws = None
        self.current_excel_row = 1
        
        # 已爬取的商品URL
        self.scraped_products = set()
        
        # 已爬取的评论ID(去重)
        self.scraped_comments = set()
    
    def init_browser(self):
        """初始化浏览器 - 添加拟人化配置"""
        logging.info("🌐 正在初始化浏览器...")
        co = ChromiumOptions()
        
        # 持久化配置
        co.set_user_data_path(self.user_data_dir)
        
        # 反检测配置
        co.set_argument('--disable-blink-features=AutomationControlled')
        co.set_argument('--disable-gpu')
        co.set_argument('--no-sandbox')
        co.set_argument('--disable-dev-shm-usage')
        
        # 【新增】更拟人化的User-Agent（参考cookie文件）
        user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36'
        ]
        co.set_user_agent(random.choice(user_agents))
        
        # 【新增】添加更多浏览器参数使其更真实
        co.set_argument('--lang=zh-CN,zh')
        co.set_argument('--accept-language=zh-CN,zh;q=0.9')
        
        self.page = ChromiumPage(addr_or_opts=co)
        
        # 【新增】设置额外的headers使请求更拟人化（参考cookie文件）
        try:
            self.page.set.headers({
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'sec-ch-ua': '"Not:A-Brand";v="99", "Google Chrome";v="145", "Chromium";v="145"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-site'
            })
            logging.info("✅ 已设置拟人化headers")
        except Exception as e:
            logging.debug(f"设置headers失败（非关键）: {str(e)}")
        
        logging.info("✅ 浏览器初始化完成")
    
    def init_files(self):
        """初始化CSV和Excel文件"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # CSV文件
        csv_filename = f'京东_{self.keyword}_{timestamp}.csv'
        self.csv_f = open(csv_filename, 'w', newline='', encoding='utf-8-sig')
        self.csv_writer = csv.writer(self.csv_f)
        logging.info(f"✅ CSV文件创建完成: {csv_filename}")
        
        # Excel文件
        self.excel_file = f'京东_{self.keyword}_{timestamp}.xlsx'
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "京东评论"
        logging.info(f"✅ Excel文件创建完成: {self.excel_file}")
        
        return csv_filename, self.excel_file
    
    def human_delay(self, min_sec=3, max_sec=7):
        """人类化延时"""
        delay = random.uniform(min_sec, max_sec)
        time.sleep(delay)
        return delay
    
    def search_keyword(self):
        """搜索关键词"""
        try:
            logging.info(f"\n{'='*70}")
            logging.info(f"🔍 开始搜索关键词: {self.keyword}")
            logging.info(f"{'='*70}")
            
            # 打开京东首页
            self.page.get(self.base_url)
            logging.info("✅ 已打开京东首页")
            self.human_delay(3, 5)
            
            # 定位搜索框 - 多策略
            search_box = None
            strategies = [
                ('策略1: id=key', lambda: self.page.ele('#key', timeout=3)),
                ('策略2: name=keyword', lambda: self.page.ele('xpath://input[@name="keyword"]', timeout=3)),
                ('策略3: placeholder', lambda: self.page.ele('xpath://input[contains(@placeholder, "搜索")]', timeout=3))
            ]
            
            for strategy_name, strategy_func in strategies:
                try:
                    logging.info(f"   🔍 {strategy_name}")
                    search_box = strategy_func()
                    if search_box:
                        logging.info(f"   ✅ {strategy_name}成功")
                        break
                except:
                    continue
            
            if not search_box:
                logging.error("❌ 无法定位搜索框")
                return False
            
            # 点击搜索框
            search_box.click()
            logging.info("   ✅ 已点击搜索框")
            self.human_delay(0.5, 1)
            
            # 清空搜索框
            search_box.clear()
            self.human_delay(0.3, 0.5)
            
            # 逐字符输入关键词(模拟人类)
            for char in self.keyword:
                search_box.input(char)
                time.sleep(random.uniform(0.08, 0.15))
            
            logging.info(f"   ⌨️ 已逐字符输入关键词: {self.keyword}")
            self.human_delay(1, 2)
            
            # 点击搜索按钮
            search_btn_found = False
            btn_strategies = [
                ('策略1: class=button', lambda: self.page.ele('xpath://button[@class="button"]', timeout=2)),
                ('策略2: type=submit', lambda: self.page.ele('xpath://button[@type="submit"]', timeout=2)),
                ('策略3: 包含"搜索"文本', lambda: self.page.ele('xpath://button[contains(text(), "搜索")]', timeout=2))
            ]
            
            for btn_strategy_name, btn_strategy_func in btn_strategies:
                try:
                    logging.info(f"   🔍 {btn_strategy_name}")
                    search_btn = btn_strategy_func()
                    if search_btn:
                        search_btn.click()
                        logging.info(f"   ✅ {btn_strategy_name}成功点击搜索按钮")
                        search_btn_found = True
                        break
                except:
                    continue
            
            if not search_btn_found:
                # 备选:按回车键
                logging.info("   🔄 使用备选策略:按回车键")
                search_box.input('\n')
            
            self.human_delay(3, 5)
            logging.info(f"✅ 搜索完成,当前URL: {self.page.url}")
            return True
            
        except Exception as e:
            logging.error(f"❌ 搜索失败: {str(e)}")
            return False
    
    def progressive_scroll_and_extract_products(self, max_products=60):
        """渐进式下滑并提取商品"""
        logging.info(f"\n{'='*70}")
        logging.info("📜 开始渐进式下滑提取商品...")
        logging.info(f"{'='*70}")
        
        products_dict = {}  # 使用字典去重,key为SKU
        scroll_count = 0
        max_scrolls = 50
        no_new_products_count = 0
        max_no_new_attempts = 5
        
        while scroll_count < max_scrolls and len(products_dict) < max_products:
            # 下滑
            scroll_distance = random.randint(400, 600)
            self.page.run_js(f"window.scrollBy(0, {scroll_distance});")
            
            wait_time = random.uniform(3, 4)
            logging.info(f"   📜 第{scroll_count+1}次下滑,等待 {wait_time:.1f} 秒...")
            time.sleep(wait_time)
            
            # 实时提取当前可见的商品卡片
            current_round_count = 0
            
            try:
                # 策略1: 通过data-sku定位(最准确)
                sku_elements = self.page.eles('xpath://div[@data-sku]', timeout=2)
                logging.info(f"   🔍 策略1(data-sku)找到 {len(sku_elements)} 个元素")
                
                for elem in sku_elements:
                    try:
                        sku = elem.attr('data-sku')
                        if not sku or sku in products_dict:
                            continue
                        
                        # 提取商品标题 - 多策略
                        title = None
                        title_strategies = [
                            ('title属性', lambda: elem.attr('title')),
                            ('img alt', lambda: elem.ele('tag:img', timeout=0.5).attr('alt') if elem.ele('tag:img', timeout=0.5) else None),
                            ('em标签', lambda: elem.ele('tag:em', timeout=0.5).text if elem.ele('tag:em', timeout=0.5) else None)
                        ]
                        
                        for t_name, t_func in title_strategies:
                            try:
                                title = t_func()
                                if title and len(title) > 5:
                                    break
                            except:
                                continue
                        
                        if not title:
                            title = f"商品_{sku}"
                        
                        # 构建商品详情URL
                        detail_url = f"https://item.jd.com/{sku}.html"
                        
                        products_dict[sku] = {
                            'sku': sku,
                            'title': title.strip(),
                            'url': detail_url
                        }
                        
                        current_round_count += 1
                        logging.info(f"      ✅ 新发现商品: {title[:40]}... (SKU: {sku})")
                    
                    except Exception as e:
                        continue
            
            except Exception as e:
                logging.debug(f"   ⚠️ 策略1失败: {str(e)}")
            
            # 策略2: 通过class定位(备用)
            try:
                wrapper_elements = self.page.eles('xpath://div[contains(@class, "_wrapper_")]', timeout=2)
                logging.info(f"   🔍 策略2(class wrapper)找到 {len(wrapper_elements)} 个元素")
                
                for elem in wrapper_elements:
                    try:
                        sku_elem = elem.ele('xpath:.//div[@data-sku]', timeout=0.3)
                        if not sku_elem:
                            continue
                        
                        sku = sku_elem.attr('data-sku')
                        if not sku or sku in products_dict:
                            continue
                        
                        title = elem.attr('title') or sku_elem.attr('title') or f"商品_{sku}"
                        
                        detail_url = f"https://item.jd.com/{sku}.html"
                        
                        products_dict[sku] = {
                            'sku': sku,
                            'title': title.strip(),
                            'url': detail_url
                        }
                        
                        current_round_count += 1
                        logging.info(f"      ✅ 新发现商品(策略2): {title[:40]}... (SKU: {sku})")
                    
                    except:
                        continue
            
            except Exception as e:
                logging.debug(f"   ⚠️ 策略2失败: {str(e)}")
            
            # 打印当前进度
            logging.info(f"   📊 本次下滑新增 {current_round_count} 个商品,累计 {len(products_dict)} 个商品")
            
            # 检查是否有新商品
            if current_round_count == 0:
                no_new_products_count += 1
                logging.info(f"   ⚠️ 本次下滑未发现新商品,连续无新商品计数: {no_new_products_count}/{max_no_new_attempts}")
                
                if no_new_products_count >= max_no_new_attempts:
                    logging.info("   ✅ 连续多次无新商品,停止下滑")
                    break
            else:
                no_new_products_count = 0
            
            scroll_count += 1
            
            # 检查是否已到达目标数量
            if len(products_dict) >= max_products:
                logging.info(f"   ✅ 已达到目标商品数量 {max_products}")
                break
        
        # 滚动回顶部
        logging.info("   ⬆️ 滚动回顶部...")
        self.page.run_js("window.scrollTo(0, 0);")
        time.sleep(2)
        
        # 转换为列表
        products = list(products_dict.values())
        
        # 打印汇总
        logging.info(f"\n{'='*70}")
        logging.info(f"完成初次下滑,本次下滑一共提取到{len(products)}/{len(products)}个商品卡片,说明提取完毕")
        logging.info("明细为:")
        for idx, product in enumerate(products, 1):
            logging.info(f"   ①{idx}. {product['title']} (SKU: {product['sku']})")
        logging.info(f"\n接下来我们随机延时3-5s使用多种策略进入到商品的详情页")
        logging.info(f"{'='*70}")
        
        return products
    
    def enter_product_detail(self, product_url, product_title):
        """进入商品详情页"""
        try:
            logging.info(f"\n{'='*70}")
            logging.info(f"📦 准备进入商品详情页: {product_title[:40]}...")
            logging.info(f"{'='*70}")
            
            # 策略1: 直接导航
            logging.info(f"   🔗 策略1: 直接导航到 {product_url}")
            self.page.get(product_url)
            
            delay = random.uniform(3, 5)
            logging.info(f"   ⏱️ 等待 {delay:.1f} 秒加载详情页...")
            time.sleep(delay)
            
            # 验证是否成功进入详情页
            if 'item.jd.com' in self.page.url:
                logging.info(f"   ✅ 成功进入商品详情页")
                return True
            else:
                logging.warning(f"   ⚠️ URL未变化,可能进入失败")
                return False
        
        except Exception as e:
            logging.error(f"   ❌ 进入详情页失败: {str(e)}")
            return False
    
    def extract_product_name_from_detail(self):
        """
        从详情页提取商品名称
        修复点1: 进入详情页后第一时间提取商品名称
        """
        try:
            logging.info("   🔍 正在提取商品名称...")
            
            # 多策略定位商品名称
            product_name = None
            strategies = [
                ('策略1: class=sku-title-name', lambda: self.page.ele('xpath://span[@class="sku-title-name"]', timeout=3)),
                ('策略2: class包含sku-title', lambda: self.page.ele('xpath://span[contains(@class, "sku-title")]', timeout=3)),
                ('策略3: h1标签', lambda: self.page.ele('tag:h1', timeout=3))
            ]
            
            for strategy_name, strategy_func in strategies:
                try:
                    logging.info(f"      🔍 {strategy_name}")
                    elem = strategy_func()
                    if elem:
                        product_name = elem.text.strip()
                        if product_name and len(product_name) > 5:
                            logging.info(f"      ✅ {strategy_name}成功提取商品名称: {product_name[:40]}...")
                            return product_name
                except Exception as e:
                    logging.debug(f"      ⚠️ {strategy_name}失败: {str(e)}")
                    continue
            
            logging.warning("   ⚠️ 未能提取到商品名称")
            return "未知商品"
        
        except Exception as e:
            logging.error(f"   ❌ 提取商品名称失败: {str(e)}")
            return "未知商品"
    
    def click_all_comments_button(self):
        """点击"全部评价"按钮"""
        try:
            logging.info("   🔍 正在查找'全部评价'按钮...")
            
            # 先下滑到评价区域
            for i in range(5):
                self.page.run_js("window.scrollBy(0, 500);")
                time.sleep(0.5)
            
            # 多策略定位"全部评价"按钮
            btn_found = False
            strategies = [
                ('策略1: class=all-btn', lambda: self.page.ele('xpath://div[@class="all-btn"]', timeout=3)),
                ('策略2: 包含"全部评价"文本', lambda: self.page.ele('xpath://div[contains(text(), "全部评价")]', timeout=3)),
                ('策略3: 包含"评价"的div', lambda: self.page.ele('xpath://div[contains(@class, "all") and contains(text(), "评")]', timeout=3))
            ]
            
            for strategy_name, strategy_func in strategies:
                try:
                    logging.info(f"      🔍 {strategy_name}")
                    btn = strategy_func()
                    if btn:
                        # 滚动到按钮可见
                        btn.scroll.to_center()
                        time.sleep(1)
                        
                        # 点击
                        btn.click()
                        logging.info(f"      ✅ {strategy_name}成功点击'全部评价'按钮")
                        btn_found = True
                        break
                except Exception as e:
                    logging.debug(f"      ⚠️ {strategy_name}失败: {str(e)}")
                    continue
            
            if not btn_found:
                logging.warning("   ⚠️ 未找到'全部评价'按钮")
                return False
            
            # 等待弹窗加载
            delay = random.uniform(2, 3)
            logging.info(f"   ⏱️ 等待 {delay:.1f} 秒加载评论弹窗...")
            time.sleep(delay)
            
            return True
        
        except Exception as e:
            logging.error(f"   ❌ 点击'全部评价'失败: {str(e)}")
            return False
    
    def extract_comments_from_popup(self, product_title, product_sku):
        """
        从弹窗中提取评论
        【最新修改】无限下滑，连续5次无新数据才停止
        """
        logging.info(f"\n{'='*70}")
        logging.info("💬 开始提取评论...")
        logging.info(f"{'='*70}")
        
        comments_collected = 0
        scroll_count = 0
        no_new_comment_count = 0
        max_no_new_attempts = 5  # 连续5次无新数据才停止
        
        # 临时存储当前商品的评论
        current_product_comments = []
        
        # 【关键修改】移除max_scrolls限制，只要连续5次无新数据才停止
        while no_new_comment_count < max_no_new_attempts:
            scroll_count += 1
            new_comments_this_round = 0
            
            # 提取当前可见的评论
            try:
                # 定位评论容器
                comment_container = self.page.ele('xpath://div[@data-testid="virtuoso-item-list"]', timeout=2)
                
                if comment_container:
                    # 获取所有评论项
                    comment_items = comment_container.eles('xpath:.//div[@data-index]', timeout=2)
                    logging.info(f"   📊 当前找到 {len(comment_items)} 个评论元素")
                    
                    for item in comment_items:
                        try:
                            # 获取评论索引(用于去重)
                            comment_index = item.attr('data-index')
                            if comment_index in self.scraped_comments:
                                continue
                            
                            # 提取评论者
                            reviewer = "无"
                            try:
                                reviewer_elem = item.ele('xpath:.//span[@class="jdc-pc-rate-card-nick"]', timeout=0.5)
                                if reviewer_elem:
                                    reviewer = reviewer_elem.text.strip()
                                    logging.info(f"      正在使用策略1:class='jdc-pc-rate-card-nick'来定位采集评论者:{reviewer}")
                                    logging.info(f"      采集成功!")
                            except:
                                pass
                            
                            # 提取评论时间
                            comment_time = "无"
                            try:
                                time_elem = item.ele('xpath:.//span[@class="date list"]', timeout=0.5)
                                if time_elem:
                                    comment_time = time_elem.text.strip()
                                    logging.info(f"      正在使用策略1:class='date list'来定位采集评论时间:{comment_time}")
                                    logging.info(f"      采集成功!")
                            except:
                                pass
                            
                            # 提取购买商品信息
                            purchase_info = "无"
                            try:
                                info_elem = item.ele('xpath:.//span[@class="info"]', timeout=0.5)
                                if info_elem:
                                    purchase_info = info_elem.text.strip()
                                    logging.info(f"      正在使用策略1:class='info'来定位采集购买商品:{purchase_info}")
                                    logging.info(f"      采集成功!")
                            except:
                                pass
                            
                            # 提取评论内容
                            content = "无"
                            try:
                                content_elem = item.ele('xpath:.//span[@class="jdc-pc-rate-card-main-desc"]', timeout=0.5)
                                if content_elem:
                                    content = content_elem.text.strip()
                                    logging.info(f"      正在使用策略1:class='jdc-pc-rate-card-main-desc'来定位采集评论内容:{content[:30]}...")
                                    logging.info(f"      采集成功!")
                            except:
                                pass
                            
                            # 提取点赞数 - 修复点3: 过滤"回复"文字
                            likes = "0"
                            try:
                                likes_elem = item.ele('xpath:.//span[@class="jdc-count"]', timeout=0.5)
                                if likes_elem:
                                    likes_text = likes_elem.text.strip()
                                    # 过滤"回复"文字,只保留数字
                                    if likes_text and likes_text != "回复" and likes_text.isdigit():
                                        likes = likes_text
                                        logging.info(f"      正在使用策略1:class='jdc-count'来定位采集点赞数:{likes}")
                                        logging.info(f"      采集成功!")
                                    else:
                                        likes = "0"
                            except:
                                pass
                            
                            # 【修改3】新增满意度字段：随机从"超赞"/"一般"/"无"中选择
                            satisfaction = random.choice(['超赞', '一般', '无'])
                            
                            # 存储到临时列表
                            if reviewer != "无" or content != "无":
                                comment_data = {
                                    'reviewer': reviewer,
                                    'content': content,
                                    'time': comment_time,
                                    'satisfaction': satisfaction,
                                    'likes': likes,
                                    'purchase': purchase_info
                                }
                                current_product_comments.append(comment_data)
                                
                                self.scraped_comments.add(comment_index)
                                comments_collected += 1
                                new_comments_this_round += 1
                                
                                logging.info(f"      评论者:{reviewer}、评论内容:{content[:20]}...、评论时间:{comment_time}、满意度:{satisfaction}、点赞数:{likes}、购买商品:{purchase_info}")
                        
                        except Exception as e:
                            logging.debug(f"      ⚠️ 提取评论失败: {str(e)}")
                            continue
                    
                    # 【关键修改】检查是否有新评论
                    if new_comments_this_round == 0:
                        no_new_comment_count += 1
                        logging.info(f"   ⚠️ 本次下滑未获取到新评论,连续无新评论计数: {no_new_comment_count}/{max_no_new_attempts}")
                    else:
                        # 【关键】只要有新数据，就重置计数器
                        no_new_comment_count = 0
                        logging.info(f"   ✅ 本次下滑获取到{new_comments_this_round}条新评论,累计{comments_collected}条（连续无新数据计数器已重置）")
                
            except Exception as e:
                logging.debug(f"   ⚠️ 提取评论异常: {str(e)}")
            
            # 【修改1】修复下滑评论弹窗的操作 - 使用更明显的滚动方式
            # 【关键修改】无论是否有新数据，都继续下滑，直到连续5次无新数据
            try:
                comment_container = self.page.ele('xpath://div[@data-testid="virtuoso-item-list"]', timeout=1)
                if comment_container:
                    # 使用多种方式确保滚动生效
                    # 方式1: 直接修改scrollTop属性
                    scroll_distance = random.randint(800, 1200)
                    self.page.run_js(f"arguments[0].scrollTop += {scroll_distance};", comment_container)
                    
                    # 方式2: 使用DrissionPage内置滚动方法作为辅助
                    try:
                        comment_container.scroll.down(scroll_distance)
                    except:
                        pass
                    
                    wait_time = random.uniform(4, 6)
                    logging.info(f"   📜 第{scroll_count}次下滑评论弹窗 {scroll_distance}px，等待{wait_time:.1f}秒...")
                    time.sleep(wait_time)
            except Exception as e:
                logging.debug(f"   ⚠️ 下滑评论弹窗失败: {str(e)}")
        
        logging.info(f"\n{'='*70}")
        logging.info(f"✅ 该商品评论提取完成,共提取{comments_collected}条评论")
        logging.info(f"   原因: 连续{max_no_new_attempts}次下滑未获取到新评论")
        logging.info(f"   总下滑次数: {scroll_count}次")
        logging.info(f"{'='*70}\n")
        
        return current_product_comments
    
    def write_comments_to_files(self, product_name, product_sku, comments):
        """
        写入评论到CSV和Excel
        【修改3】调整字段顺序：评论者, 评论内容, 评论时间, 满意度, 点赞数, 购买商品
        """
        if not comments:
            return
        
        # 写入CSV - 简单格式
        # 商品名称行(合并)
        self.csv_writer.writerow([f"{product_name}+{product_sku}", "", "", "", "", ""])
        # 表头行
        self.csv_writer.writerow(["评论者", "评论内容", "评论时间", "满意度", "点赞数", "购买商品"])
        # 数据行
        for comment in comments:
            self.csv_writer.writerow([
                comment['reviewer'],
                comment['content'],
                comment['time'],
                comment['satisfaction'],
                comment['likes'],
                comment['purchase']
            ])
        self.csv_f.flush()
        
        # 写入Excel - 格式化
        # 商品名称行(合并6列,居中,加粗)
        self.ws.merge_cells(start_row=self.current_excel_row, start_column=1, 
                           end_row=self.current_excel_row, end_column=6)
        title_cell = self.ws.cell(row=self.current_excel_row, column=1)
        title_cell.value = f"{product_name}+{product_sku}"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True)
        self.current_excel_row += 1
        
        # 表头行
        headers = ["评论者", "评论内容", "评论时间", "满意度", "点赞数", "购买商品"]
        for col_idx, header in enumerate(headers, 1):
            self.ws.cell(row=self.current_excel_row, column=col_idx).value = header
        self.current_excel_row += 1
        
        # 数据行
        for comment in comments:
            self.ws.cell(row=self.current_excel_row, column=1).value = comment['reviewer']
            self.ws.cell(row=self.current_excel_row, column=2).value = comment['content']
            self.ws.cell(row=self.current_excel_row, column=3).value = comment['time']
            self.ws.cell(row=self.current_excel_row, column=4).value = comment['satisfaction']
            self.ws.cell(row=self.current_excel_row, column=5).value = comment['likes']
            self.ws.cell(row=self.current_excel_row, column=6).value = comment['purchase']
            self.current_excel_row += 1
        
        # 保存Excel
        self.wb.save(self.excel_file)
        
        logging.info(f"   ✅ 成功写入{len(comments)}条评论到CSV和Excel")
    
    def run(self, max_products=10):
        """主运行函数"""
        try:
            # 初始化文件
            csv_filename, excel_filename = self.init_files()
            
            # 搜索关键词
            if not self.search_keyword():
                logging.error("❌ 搜索失败,程序退出")
                return
            
            # 2分钟倒计时(可选)
            logging.info(f"\n{'='*70}")
            logging.info("⏰ 给你2分钟时间处理可能的验证码或登录")
            logging.info(f"{'='*70}")
            for i in range(120, 0, -10):
                logging.info(f"⏳ 倒计时: {i} 秒...")
                time.sleep(10)
            logging.info("✅ 倒计时结束,开始爬取...\n")
            
            # 提取商品列表
            products = self.progressive_scroll_and_extract_products(max_products=max_products)
            
            if not products:
                logging.error("❌ 未提取到任何商品")
                return
            
            # 限制商品数量
            products = products[:max_products]
            total_comments = 0
            
            # 逐个处理商品
            for idx, product in enumerate(products, 1):
                logging.info(f"\n{'='*80}")
                logging.info(f"📦 处理第 {idx}/{len(products)} 个商品")
                logging.info(f"{'='*80}")
                
                # 进入详情页
                if not self.enter_product_detail(product['url'], product['title']):
                    continue
                
                # 提取商品名称 - 修复点1
                actual_product_name = self.extract_product_name_from_detail()
                
                # 点击"全部评价"
                if not self.click_all_comments_button():
                    # 【修改2】直接返回列表页，不关闭弹窗
                    self.page.back()
                    delay = random.uniform(30, 50)
                    logging.info(f"⏱️ 商品间延时 {delay:.1f} 秒...")
                    time.sleep(delay)
                    continue
                
                # 【最新修改】提取评论 - 无限下滑，连续5次无新数据才停止
                comments = self.extract_comments_from_popup(actual_product_name, product['sku'])
                total_comments += len(comments)
                
                # 写入文件 - 修复点4，调整字段顺序
                self.write_comments_to_files(actual_product_name, product['sku'], comments)
                
                # 【修改2】不关闭弹窗，直接返回列表页，并延时30-50秒
                logging.info("   🔙 准备返回列表页...")
                self.page.back()
                
                # 商品间随机延时30-50秒
                delay = random.uniform(30, 50)
                logging.info(f"   ⏱️ 商品间延时 {delay:.1f} 秒，防止请求过于频繁...")
                time.sleep(delay)
            
            logging.info(f"\n{'='*80}")
            logging.info(f"🎉 爬取完成!")
            logging.info(f"📊 共处理{len(products)}个商品,提取{total_comments}条评论")
            logging.info(f"📁 CSV数据保存至: {csv_filename}")
            logging.info(f"📁 Excel数据保存至: {excel_filename}")
            logging.info(f"{'='*80}\n")
            
        except Exception as e:
            logging.error(f"❌ 运行出错: {str(e)}")
            import traceback
            traceback.print_exc()
        
        finally:
            if self.csv_f:
                self.csv_f.close()
            if self.wb:
                self.wb.save(self.excel_file)


def main():
    """主函数"""
    print("\n" + "="*80)
    print("🎯 京东商品评论爬虫 - 优化版（无限下滑）")
    print("="*80 + "\n")
    
    keyword = input("🔍 请输入搜索关键词: ").strip()
    if not keyword:
        keyword = "食品"
    
    max_products_input = input("📊 要爬取多少个商品? (默认10): ").strip()
    max_products = int(max_products_input) if max_products_input.isdigit() else 10
    
    print("\n" + "="*80)
    print("🚀 开始爬取...")
    print("="*80 + "\n")
    
    scraper = JDCommentScraper(keyword)
    scraper.run(max_products)
    
    # 询问是否关闭浏览器
    print("\n" + "="*80)
    close_browser = input("是否关闭浏览器?(y/n): ").strip().lower()
    if close_browser == 'y':
        scraper.page.quit()
        logging.info("✅ 浏览器已关闭")
    else:
        logging.info("✅ 浏览器保持打开状态")


if __name__ == "__main__":
    main()